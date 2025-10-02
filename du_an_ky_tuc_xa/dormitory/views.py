# dormitory/views.py
from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from .models import Room, Building, Contract, Student

def home(request):
    total_rooms = Room.objects.count()
    available_rooms = Room.objects.filter(status='available').count()
    total_buildings = Building.objects.count()
    active_contracts = Contract.objects.filter(status='active').count()
    
    context = {
        'total_rooms': total_rooms,
        'available_rooms': available_rooms,
        'total_buildings': total_buildings,
        'active_contracts': active_contracts,
    }
    return render(request, 'dormitory/home.html', context)

@login_required
# def dashboard(request):
#     # Dashboard cho quản lý
#     if request.user.user_type not in ['manager', 'staff']:
#         messages.error(request, "Bạn không có quyền truy cập trang quản lý")
#         return redirect('home')
    
#     context = {
#         'stats': {
#             'total_rooms': Room.objects.count(),
#             'available_rooms': Room.objects.filter(status='available').count(),
#             'active_contracts': Contract.objects.filter(status='active').count(),
#         }
#     }
#     return render(request, 'dormitory/dashboard.html', context)

@login_required
def student_dashboard(request):
    # Dashboard cho sinh viên
    if request.user.user_type != 'student':
        messages.error(request, "Bạn không có quyền truy cập trang này")
        return redirect('home')
    
    try:
        student = Student.objects.get(user=request.user)
        current_contract = Contract.objects.filter(
            student=student, 
            status='active'
        ).first()
        
        # Lấy danh sách phòng trống
        available_rooms = Room.objects.filter(
            status='available'
        ).select_related('building', 'room_type')
        
        context = {
            'student': student,
            'current_contract': current_contract,
            'available_rooms': available_rooms,
        }
        return render(request, 'dormitory/student_dashboard.html', context)
        
    except Student.DoesNotExist:
        return redirect('complete_profile')

@login_required
def complete_profile(request):
    # Hoàn thiện hồ sơ sinh viên
    if request.user.user_type != 'student':
        return redirect('home')
    
    if request.method == 'POST':
        student_id = request.POST['student_id']
        university = request.POST['university']
        faculty = request.POST['faculty']
        course = request.POST['course']
        
        # Kiểm tra xem student_id đã tồn tại chưa
        if Student.objects.filter(student_id=student_id).exists():
            messages.error(request, "Mã sinh viên đã tồn tại!")
            return render(request, 'dormitory/complete_profile.html')
        
        Student.objects.create(
            user=request.user,
            student_id=student_id,
            university=university,
            faculty=faculty,
            course=course
        )
        
        messages.success(request, "Hoàn tất hồ sơ thành công!")
        return redirect('student_dashboard')
    
    return render(request, 'dormitory/complete_profile.html')

def dashboard(request):
    # Chỉ cho phép manager/staff truy cập
    # if request.user.user_type not in ['manager', 'staff']:
    #     return render(request, 'errors/access_denied.html')
    
    # Thống kê tổng quan
    total_buildings = Building.objects.count()
    total_rooms = Room.objects.count()
    available_rooms = Room.objects.filter(status='available').count()
    occupied_rooms = Room.objects.filter(status='occupied').count()
    total_students = Student.objects.count()
    active_contracts = Contract.objects.filter(status='active').count()
    
    # Tính tỷ lệ lấp đầy
    if total_rooms > 0:
        occupancy_percentage = (occupied_rooms / total_rooms) * 100
    else:
        occupancy_percentage = 0
    
    # Phòng sắp hết hợp đồng (trong 30 ngày tới)
    from datetime import date, timedelta
    upcoming_expiry = Contract.objects.filter(
        status='active',
        end_date__lte=date.today() + timedelta(days=30)
    ).count()
    
    # THÊM PHẦN NÀY: THÔNG BÁO THANH TOÁN
    from payment.models import Payment
    from django.utils import timezone
    
    today = timezone.now().date()
    
    # Hóa đơn quá hạn
    overdue_payments = Payment.objects.filter(
        status='pending', 
        due_date__lt=today
    ).select_related('contract__student', 'contract__room')[:5]  # 5 cái gần nhất
    
    # Hóa đơn sắp đến hạn (7 ngày tới)
    upcoming_payments = Payment.objects.filter(
        status='pending',
        due_date__range=[today, today + timedelta(days=7)]
    ).select_related('contract__student', 'contract__room')[:5]
    
    # Thống kê thanh toán
    total_pending_payments = Payment.objects.filter(status='pending').count()
    total_overdue_payments = Payment.objects.filter(status='pending', due_date__lt=today).count()
    
    context = {
        'stats': {
            'total_buildings': total_buildings,
            'total_rooms': total_rooms,
            'available_rooms': available_rooms,
            'occupied_rooms': occupied_rooms,
            'occupancy_percentage': occupancy_percentage,
            'total_students': total_students,
            'active_contracts': active_contracts,
            'upcoming_expiry': upcoming_expiry,
            # THÊM THỐNG KÊ THANH TOÁN
            'total_pending_payments': total_pending_payments,
            'total_overdue_payments': total_overdue_payments,
        },
        # THÊM THÔNG BÁO
        'overdue_payments': overdue_payments,
        'upcoming_payments': upcoming_payments,
        'today': today,
    }
    return render(request, 'dormitory/dashboard.html', context)

# dormitory/views.py - THÊM CUỐI FILE
from django.shortcuts import render, get_object_or_404, redirect
from .forms import RoomForm
from django.db.models import Q
# dormitory/views.py - SỬA room_list
from django.core.paginator import Paginator

def room_list(request):
    """Danh sách phòng với tìm kiếm và phân trang"""
    rooms = Room.objects.select_related('building', 'room_type').all()
    
    # Tìm kiếm
    search_query = request.GET.get('search', '')
    if search_query:
        rooms = rooms.filter(
            Q(room_number__icontains=search_query) |
            Q(building__name__icontains=search_query) |
            Q(room_type__name__icontains=search_query)
        )
    
    # Phân trang - 10 items per page
    paginator = Paginator(rooms, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    return render(request, 'dormitory/room_list.html', {
        'page_obj': page_obj,
        'search_query': search_query
    })


def room_create(request):
    """Thêm phòng mới"""
    if request.method == 'POST':
        form = RoomForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('room_list')
    else:
        form = RoomForm()
    return render(request, 'dormitory/room_form.html', {'form': form})

def room_update(request, pk):
    """Sửa phòng"""
    room = get_object_or_404(Room, pk=pk)
    if request.method == 'POST':
        form = RoomForm(request.POST, instance=room)
        if form.is_valid():
            form.save()
            return redirect('room_list')
    else:
        form = RoomForm(instance=room)
    return render(request, 'dormitory/room_form.html', {'form': form})

def room_delete(request, pk):
    """Xóa phòng"""
    room = get_object_or_404(Room, pk=pk)
    if request.method == 'POST':
        room.delete()
        return redirect('room_list')
    return render(request, 'dormitory/room_confirm_delete.html', {'room': room})

# dormitory/views.py - THÊM CUỐI FILE
from .forms import RoomForm, BuildingForm

def building_list(request):
    """Danh sách tòa nhà với tìm kiếm và phân trang"""
    buildings = Building.objects.all()
    
    search_query = request.GET.get('search', '')
    if search_query:
        buildings = buildings.filter(
            Q(name__icontains=search_query) |
            Q(address__icontains=search_query)
        )
    
    # Phân trang - 10 items per page
    paginator = Paginator(buildings, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    return render(request, 'dormitory/building_list.html', {
        'page_obj': page_obj,
        'search_query': search_query
    })
def building_create(request):
    """Thêm tòa nhà mới"""
    if request.method == 'POST':
        form = BuildingForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('building_list')
    else:
        form = BuildingForm()
    return render(request, 'dormitory/building_form.html', {'form': form})

def building_update(request, pk):
    """Sửa tòa nhà"""
    building = get_object_or_404(Building, pk=pk)
    if request.method == 'POST':
        form = BuildingForm(request.POST, instance=building)
        if form.is_valid():
            form.save()
            return redirect('building_list')
    else:
        form = BuildingForm(instance=building)
    return render(request, 'dormitory/building_form.html', {'form': form})

def building_delete(request, pk):
    """Xóa tòa nhà"""
    building = get_object_or_404(Building, pk=pk)
    if request.method == 'POST':
        building.delete()
        return redirect('building_list')
    return render(request, 'dormitory/building_confirm_delete.html', {'building': building})

# dormitory/views.py - THÊM CUỐI FILE
from .forms import RoomForm, BuildingForm, StudentForm

def student_list(request):
    """Danh sách sinh viên với tìm kiếm và phân trang"""
    students = Student.objects.select_related('user').all()
    
    search_query = request.GET.get('search', '')
    if search_query:
        students = students.filter(
            Q(student_id__icontains=search_query) |
            Q(full_name__icontains=search_query) |
            Q(user__username__icontains=search_query) |
            Q(university__icontains=search_query) |
            Q(faculty__icontains=search_query)
        )
    
    # Phân trang - 10 items per page
    paginator = Paginator(students, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    return render(request, 'dormitory/student_list.html', {
        'page_obj': page_obj,
        'search_query': search_query
    })
def student_create(request):
    """Thêm sinh viên mới"""
    if request.method == 'POST':
        form = StudentForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('student_list')
    else:
        form = StudentForm()
    return render(request, 'dormitory/student_form.html', {'form': form})

def student_update(request, pk):
    """Sửa sinh viên"""
    student = get_object_or_404(Student, pk=pk)
    if request.method == 'POST':
        form = StudentForm(request.POST, instance=student)
        if form.is_valid():
            form.save()
            return redirect('student_list')
    else:
        form = StudentForm(instance=student)
    return render(request, 'dormitory/student_form.html', {'form': form})

def student_delete(request, pk):
    """Xóa sinh viên"""
    student = get_object_or_404(Student, pk=pk)
    if request.method == 'POST':
        student.delete()
        return redirect('student_list')
    return render(request, 'dormitory/student_confirm_delete.html', {'student': student})


# dormitory/views.py - THÊM CUỐI FILE
from .forms import RoomForm, BuildingForm, StudentForm, ContractForm

def contract_list(request):
    """Danh sách hợp đồng với tìm kiếm và phân trang"""
    contracts = Contract.objects.select_related('student', 'room').all()
    
    search_query = request.GET.get('search', '')
    if search_query:
        contracts = contracts.filter(
            Q(contract_number__icontains=search_query) |
            Q(student__student_id__icontains=search_query) |
            Q(student__full_name__icontains=search_query) |
            Q(room__room_number__icontains=search_query) |
            Q(room__building__name__icontains=search_query)
        )
    
    # Phân trang - 10 items per page
    paginator = Paginator(contracts, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    return render(request, 'dormitory/contract_list.html', {
        'page_obj': page_obj,
        'search_query': search_query
    })
def contract_create(request):
    """Thêm hợp đồng mới"""
    if request.method == 'POST':
        form = ContractForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('contract_list')
    else:
        form = ContractForm()
    return render(request, 'dormitory/contract_form.html', {'form': form})

def contract_update(request, pk):
    """Sửa hợp đồng"""
    contract = get_object_or_404(Contract, pk=pk)
    if request.method == 'POST':
        form = ContractForm(request.POST, instance=contract)
        if form.is_valid():
            form.save()
            return redirect('contract_list')
    else:
        form = ContractForm(instance=contract)
    return render(request, 'dormitory/contract_form.html', {'form': form})

def contract_delete(request, pk):
    """Xóa hợp đồng"""
    contract = get_object_or_404(Contract, pk=pk)
    if request.method == 'POST':
        contract.delete()
        return redirect('contract_list')
    return render(request, 'dormitory/contract_confirm_delete.html', {'contract': contract})

# dormitory/views.py
def reports(request):
    """Trang báo cáo thống kê"""
    # Thống kê phòng
    total_rooms = Room.objects.count()
    occupied_rooms = Room.objects.filter(status='occupied').count()
    
    # Tính tỷ lệ lấp đầy
    if total_rooms > 0:
        occupancy_percentage = (occupied_rooms / total_rooms) * 100
    else:
        occupancy_percentage = 0
    
    room_stats = {
        'total': total_rooms,
        'available': Room.objects.filter(status='available').count(),
        'occupied': occupied_rooms,
        'maintenance': Room.objects.filter(status='maintenance').count(),
        'occupancy_percentage': occupancy_percentage,  # Thêm tỷ lệ phần trăm
    }
    
    # Thống kê hợp đồng
    from datetime import date, timedelta
    contract_stats = {
        'active': Contract.objects.filter(status='active').count(),
        'expired': Contract.objects.filter(status='expired').count(),
        'upcoming_expiry': Contract.objects.filter(
            status='active',
            end_date__lte=date.today() + timedelta(days=30)
        ).count(),
    }
    
    # Thống kê theo tòa nhà
    building_stats = []
    for building in Building.objects.all():
        building_total = building.room_set.count()
        building_occupied = building.room_set.filter(status='occupied').count()
        
        # Tính tỷ lệ lấp đầy cho từng tòa nhà
        if building_total > 0:
            building_occupancy_rate = (building_occupied / building_total) * 100
        else:
            building_occupancy_rate = 0
        
        building_stats.append({
            'name': building.name,
            'total_rooms': building_total,
            'occupied_rooms': building_occupied,
            'occupancy_rate': building_occupancy_rate,  # Tỷ lệ phần trăm
            'available_rooms': building_total - building_occupied,  # Phòng trống
        })
    
    context = {
        'room_stats': room_stats,
        'contract_stats': contract_stats,
        'building_stats': building_stats,
    }
    return render(request, 'dormitory/reports.html', context)


import io
from django.http import HttpResponse
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter
from openpyxl import Workbook
from django.utils import timezone
def export_rooms_pdf(request):
    """Xuất danh sách phòng PDF"""
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="danh_sach_phong.pdf"'
    
    buffer = io.BytesIO()
    p = canvas.Canvas(buffer, pagesize=letter)
    
    # Tiêu đề
    p.setFont("Helvetica-Bold", 16)
    p.drawString(100, 750, "DANH SÁCH PHÒNG - KÝ TÚC XÁ")
    p.setFont("Helvetica", 10)
    p.drawString(100, 730, f"Ngày xuất: {timezone.now().strftime('%d/%m/%Y %H:%M')}")
    
    # Header table
    p.setFont("Helvetica-Bold", 10)
    p.drawString(50, 700, "Mã phòng")
    p.drawString(120, 700, "Tòa nhà")
    p.drawString(200, 700, "Loại phòng")
    p.drawString(300, 700, "Tầng")
    p.drawString(350, 700, "Trạng thái")
    
    # Dữ liệu
    rooms = Room.objects.select_related('building', 'room_type').all()
    y = 680
    p.setFont("Helvetica", 9)
    
    for room in rooms:
        if y < 100:  # Tạo trang mới nếu hết chỗ
            p.showPage()
            y = 750
            # Vẽ header lại cho trang mới
            p.setFont("Helvetica-Bold", 10)
            p.drawString(50, 750, "Mã phòng")
            p.drawString(120, 750, "Tòa nhà")
            p.drawString(200, 750, "Loại phòng")
            p.drawString(300, 750, "Tầng")
            p.drawString(350, 750, "Trạng thái")
            p.setFont("Helvetica", 9)
            y = 730
        
        p.drawString(50, y, room.room_number)
        p.drawString(120, y, room.building.name)
        p.drawString(200, y, room.room_type.name)
        p.drawString(300, y, str(room.floor))
        p.drawString(350, y, room.get_status_display())
        y -= 20
    
    p.save()
    pdf = buffer.getvalue()
    buffer.close()
    response.write(pdf)
    return response

def export_rooms_excel(request):
    """Xuất danh sách phòng Excel"""
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="danh_sach_phong.xlsx"'
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Danh sách phòng"
    
    # Header
    headers = ['Mã phòng', 'Tòa nhà', 'Loại phòng', 'Sức chứa', 'Giá thuê', 'Tầng', 'Trạng thái']
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
    
    # Dữ liệu
    rooms = Room.objects.select_related('building', 'room_type').all()
    for row, room in enumerate(rooms, 2):
        ws.cell(row=row, column=1, value=room.room_number)
        ws.cell(row=row, column=2, value=room.building.name)
        ws.cell(row=row, column=3, value=room.room_type.name)
        ws.cell(row=row, column=4, value=room.room_type.capacity)
        ws.cell(row=row, column=5, value=float(room.room_type.price_per_month))
        ws.cell(row=row, column=6, value=room.floor)
        ws.cell(row=row, column=7, value=room.get_status_display())
    
    wb.save(response)
    return response

def export_students_excel(request):
    """Xuất danh sách sinh viên Excel"""
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="danh_sach_sinh_vien.xlsx"'
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Danh sách sinh viên"
    
    # Header
    headers = ['Mã SV', 'Họ tên', 'Ngày sinh', 'Email', 'Trường', 'Khoa', 'Khóa học']
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
    
    # Dữ liệu
    students = Student.objects.select_related('user').all()
    for row, student in enumerate(students, 2):
        ws.cell(row=row, column=1, value=student.student_id)
        ws.cell(row=row, column=2, value=student.full_name or student.user.get_full_name())
        ws.cell(row=row, column=3, value=student.date_of_birth.strftime('%d/%m/%Y') if student.date_of_birth else '')
        ws.cell(row=row, column=4, value=student.user.email)
        ws.cell(row=row, column=5, value=student.university)
        ws.cell(row=row, column=6, value=student.faculty)
        ws.cell(row=row, column=7, value=student.course)
    
    wb.save(response)
    return response

# dormitory/views.py
@login_required
def room_booking(request, room_id):
    """Đăng ký phòng cho sinh viên"""
    if request.user.user_type != 'student':
        messages.error(request, "Chỉ sinh viên mới có thể đăng ký phòng!")
        return redirect('home')
    
    room = get_object_or_404(Room, pk=room_id, status='available')
    student = request.user.student
    
    # Kiểm tra sinh viên đã có hợp đồng active chưa
    existing_contract = Contract.objects.filter(student=student, status='active').first()
    if existing_contract:
        messages.warning(request, f"Bạn đã có hợp đồng phòng {existing_contract.room.room_number}!")
        return redirect('student_dashboard')
    
    if request.method == 'POST':
        # Tạo hợp đồng mới
        from datetime import date, timedelta
        
        contract = Contract.objects.create(
            contract_number=f"CT{date.today().strftime('%Y%m%d')}{student.student_id}",
            student=student,
            room=room,
            start_date=date.today(),
            end_date=date.today() + timedelta(days=365),  # 1 năm
            deposit=room.room_type.price_per_month,  # Cọc 1 tháng
            status='active'
        )
        
        # Cập nhật trạng thái phòng
        room.status = 'occupied'
        room.save()
        
        messages.success(request, f"✅ Đã đăng ký thành công phòng {room.room_number}!")
        return redirect('student_dashboard')
    
    return render(request, 'dormitory/room_booking.html', {
        'room': room,
        'student': student
    })